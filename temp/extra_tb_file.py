import io
import os

from openpyxl import Workbook
from service.ReportService import get_all_file_record


def run():
    try:

        wb = Workbook()  
        ws = wb.active  

        
        ws['A1'] = '来源'
        ws['B1'] = '公司'
        ws['C1'] = '文件名'
        ws['D1'] = '标题'
        files = get_all_file_record()
        i = 2
        for file in files:
            last_index = file['file_path'].rfind("/")
            second_index = file['file_path'].rfind("/", 0, last_index)
            third_index = file['file_path'].rfind("/", 0, second_index)
            ticker = file['file_path'][second_index + 1:last_index]
            name = file['file_path'][last_index + 1:len(file['file_path'])]
            source = file['file_path'][third_index + 1:second_index]
            ws['A' + str(i)] = source
            ws['B' + str(i)] = ticker
            ws['C' + str(i)] = name
            ws['D' + str(i)] = file['title']
            i = i + 1
        wb.save(os.getcwd() + '/' + 'format_data.xlsx')
    except Exception as e:
        print(str(e))


if __name__ == '__main__':
    run()
